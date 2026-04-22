import os
import json
from io import BytesIO
from datetime import timedelta

from django.contrib import messages
from django.core.management import call_command
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import IntegrityError, OperationalError, ProgrammingError, transaction
from django.db.models import Count
from django.shortcuts import get_object_or_404, redirect, render
from django.http import Http404
from django.utils import timezone
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .forms import (
    LoginForm, EnhancedRegisterForm, OTPVerificationForm, Enable2FAForm,
    VoterProfileForm, CandidateForm, VoteForm, ElectionForm, LanguagePreferenceForm
)
from .models import (
    Election, Vote, CustomUser, VoterProfile, VoteLog, 
    OTPVerification, CaptchaLog, NotificationLog, Candidate, AuditLog
)
from .utils import (
    create_otp_record, verify_otp, setup_totp, verify_totp,
    send_vote_confirmation_email, create_vote_log, get_client_ip,
    get_election_stats, is_bot_activity, send_admin_notification
)


def _ensure_election_schema():
    """Add any missing Election columns before the homepage query runs."""
    from django.db import connection as db_connection

    existing_columns = set()
    if db_connection.vendor == "postgresql":
        with db_connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'voting_election'
                """
            )
            existing_columns = {row[0] for row in cursor.fetchall()}
    elif db_connection.vendor == "sqlite":
        with db_connection.cursor() as cursor:
            cursor.execute("PRAGMA table_info(voting_election)")
            existing_columns = {row[1] for row in cursor.fetchall()}

    if not existing_columns:
        return

    with db_connection.schema_editor() as schema_editor:
        for field in Election._meta.local_fields:
            if field.column not in existing_columns:
                schema_editor.add_field(Election, field)


def _ensure_sqlite_schema_for_vercel_fallback(error: OperationalError) -> bool:
    message = str(error).lower()
    is_missing_table = "no such table" in message
    using_fallback = os.getenv("VERCEL") and not (
        os.getenv("DATABASE_URL")
        or os.getenv("POSTGRES_URL_NON_POOLING")
        or os.getenv("POSTGRES_URL")
        or os.getenv("POSTGRES_PRISMA_URL")
    )
    if is_missing_table and using_fallback:
        call_command("migrate", interactive=False, verbosity=0, run_syncdb=True)
        return True
    return False


def is_admin(user):
    """Check if user is admin"""
    return user.is_staff or user.is_superuser


def get_user_ip(request):
    """Get user's IP address"""
    return get_client_ip(request)


# ============ Authentication Views ============

def home(request):
    """Home page with upcoming elections - public for all users"""
    try:
        elections = list(
            Election.objects.values("id", "title", "description").order_by("-id")
        )
    except (OperationalError, ProgrammingError):
        elections = []
    
    context = {
        'elections': elections,
        'user_language': request.user.language if request.user.is_authenticated else 'en',
    }
    return render(request, "voting/home.html", context)


def register_view(request):
    """Registration - auto-login and skip email verification"""
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = EnhancedRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Auto-login the user after registration
            login(request, user)
            messages.success(request, "Account created successfully! Welcome to E-Voting.")
            # Mark email as verified (skip OTP)
            user.email_verified = True
            user.save()
            # Log the registration
            create_vote_log(user, None, 'registered', 'User registered and auto-logged in', get_user_ip(request))
            return redirect("dashboard")
    else:
        form = EnhancedRegisterForm()

    return render(request, "voting/register.html", {"form": form})


@login_required
def verify_email(request, user_id):
    """Verify email with OTP"""
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.user.id != user.id:
        messages.error(request, "Unauthorized access")
        return redirect("home")
    
    if request.method == "POST":
        form = OTPVerificationForm(request.POST)
        if form.is_valid():
            otp_code = form.cleaned_data['otp_code']
            success, message = verify_otp(user, otp_code, 'email')
            
            if success:
                user.email_verified = True
                user.save()
                messages.success(request, "Email verified successfully!")
                return redirect("complete_voter_profile")
            else:
                messages.error(request, message)
    else:
        form = OTPVerificationForm()
    
    return render(request, "voting/verify_email.html", {"form": form, "user": user})


@login_required
def complete_voter_profile(request):
    """Complete voter profile with verification details"""
    try:
        voter_profile = request.user.voter_profile
    except VoterProfile.DoesNotExist:
        voter_profile = None
    
    if request.method == "POST":
        form = VoterProfileForm(request.POST, instance=voter_profile)
        if form.is_valid():
            profile = form.save(commit=False)
            profile.user = request.user
            profile.save()
            messages.success(request, "Voter profile updated successfully!")
            
            # Log the activity
            create_vote_log(request.user, None, 'profile_updated', 
                           f'Voter profile completed: {profile.unique_voter_id}',
                           get_user_ip(request))
            
            return redirect("setup_2fa")
    else:
        form = VoterProfileForm(instance=voter_profile)
    
    return render(request, "voting/complete_voter_profile.html", {"form": form})


@login_required
def setup_2fa(request):
    """Setup 2FA authentication"""
    if request.method == "POST":
        form = Enable2FAForm(request.POST)
        if form.is_valid():
            if form.cleaned_data.get('enable_2fa'):
                secret, qr_code = setup_totp(request.user)
                request.session['temp_otp_secret'] = secret
                return render(request, "voting/scan_qr_code.html", {
                    "qr_code": qr_code,
                    "secret": secret
                })
            else:
                messages.success(request, "Setup complete! You can enable 2FA anytime.")
                return redirect("dashboard")
    else:
        form = Enable2FAForm()
    
    return render(request, "voting/setup_2fa.html", {"form": form})


@login_required
def confirm_2fa(request):
    """Confirm 2FA setup with TOTP token"""
    if request.method == "POST":
        totp_token = request.POST.get('totp_token')
        
        if verify_totp(request.user, totp_token):
            request.user.two_fa_enabled = True
            request.user.save()
            messages.success(request, "2FA enabled successfully!")
            if 'temp_otp_secret' in request.session:
                del request.session['temp_otp_secret']
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid TOTP token. Please try again.")
    
    return render(request, "voting/confirm_2fa.html")


def login_view(request):
    """Enhanced login view with 2FA support"""
    if request.user.is_authenticated:
        return redirect("dashboard")
    
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                ip_address = get_user_ip(request)
                
                # Check for bot activity
                if is_bot_activity(ip_address):
                    messages.error(request, "Too many failed attempts. Please try again later.")
                    return render(request, "voting/login.html", {"form": form})
                
                # Check for 2FA
                if user.two_fa_enabled:
                    request.session['pre_auth_user_id'] = user.id
                    return redirect("verify_2fa")
                
                login(request, user)
                create_vote_log(user, None, 'login', f'User logged in from {ip_address}', ip_address)
                messages.success(request, "Logged in successfully!")
                return redirect("dashboard")
            else:
                messages.error(request, "Invalid username or password.")
    else:
        form = LoginForm()
    
    return render(request, "voting/login.html", {"form": form})


def verify_2fa(request):
    """Verify 2FA token during login"""
    user_id = request.session.get('pre_auth_user_id')
    if not user_id:
        return redirect("login")
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    if request.method == "POST":
        totp_token = request.POST.get('totp_token')
        
        if verify_totp(user, totp_token):
            login(request, user)
            del request.session['pre_auth_user_id']
            create_vote_log(user, None, 'login_2fa_verified', '2FA verification successful',
                           get_user_ip(request))
            messages.success(request, "2FA verified! Welcome back!")
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid 2FA token.")
    
    return render(request, "voting/verify_2fa.html", {"user": user})


# ============ Voting Views ============

@login_required
def dashboard(request):
    """User dashboard"""
    elections = list(
        Election.objects.values("id", "title", "description").order_by("-id")
    )
    
    voted_election_ids = set(
        Vote.objects.filter(user=request.user).values_list("election_id", flat=True)
    )
    
    # Check voter verification status
    try:
        voter_profile = request.user.voter_profile
        verification_status = voter_profile.verification_status
    except VoterProfile.DoesNotExist:
        verification_status = None
    
    context = {
        "elections": elections,
        "voted_election_ids": voted_election_ids,
        "voter_verification_status": verification_status,
    }
    return render(request, "voting/dashboard.html", context)


@login_required
def vote_view(request, election_id):
    """Vote in an election"""
    election = get_object_or_404(Election, id=election_id, is_active=True)
    now = timezone.now()
    ip_address = get_user_ip(request)

    if election.start_time > now or election.end_time < now:
        messages.error(request, "This election is not open for voting.")
        return redirect("dashboard")

    if not election.allow_voting:
        messages.error(request, "Voting is currently disabled for this election.")
        return redirect("dashboard")

    # Check voter verification if required
    if election.require_voter_verification:
        try:
            voter_profile = request.user.voter_profile
            if voter_profile.verification_status != 'verified':
                messages.error(request, "Your voter profile must be verified before voting.")
                return redirect("dashboard")
        except VoterProfile.DoesNotExist:
            messages.error(request, "Please complete your voter profile first.")
            return redirect("complete_voter_profile")

    if Vote.objects.filter(election=election, user=request.user).exists():
        messages.warning(request, "You have already voted in this election.")
        return redirect("results", election_id=election.id)

    if request.method == "POST":
        form = VoteForm(request.POST, election=election)
        if form.is_valid():
            candidate = form.cleaned_data["candidate"]
            try:
                with transaction.atomic():
                    vote = Vote.objects.create(
                        election=election,
                        candidate=candidate,
                        user=request.user,
                        ip_address=ip_address,
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                    
                    # Create vote log
                    create_vote_log(request.user, election, 'vote_submitted',
                                  f'Voted for {candidate.name}', ip_address)
                    
                    # Send confirmation email
                    send_vote_confirmation_email(request.user, election, candidate)
                    
            except IntegrityError:
                messages.error(request, "Duplicate vote blocked.")
                return redirect("results", election_id=election.id)

            messages.success(request, "Your vote has been submitted successfully.")
            return redirect("results", election_id=election.id)
    else:
        form = VoteForm(election=election)

    return render(request, "voting/vote.html", {"election": election, "form": form})


def results(request, election_id):
    """View election results"""
    election = get_object_or_404(Election, id=election_id)
    
    # Check if results should be published
    if not election.publish_results and not (request.user.is_authenticated and request.user.is_staff):
        messages.info(request, "Results are not yet published for this election.")
        return render(request, "voting/results_pending.html", {"election": election})

    total_votes = election.votes.count()
    candidates = election.candidates.annotate(vote_count=Count("votes")).order_by("-vote_count", "name")

    candidate_rows = []
    for candidate in candidates:
        percentage = 0
        if total_votes > 0:
            percentage = round((candidate.vote_count / total_votes) * 100, 2)
        candidate_rows.append({
            "candidate": candidate,
            "votes": candidate.vote_count,
            "percentage": percentage,
        })

    # Prepare chart data
    chart_data = json.dumps({
        'labels': [c['candidate'].name for c in candidate_rows],
        'data': [c['votes'] for c in candidate_rows],
        'backgroundColor': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40'],
    })

    context = {
        "election": election,
        "total_votes": total_votes,
        "candidate_rows": candidate_rows,
        "chart_data": chart_data,
    }
    return render(request, "voting/results.html", context)


@require_POST
@login_required
def export_results_pdf(request, election_id):
    """Export election results as PDF"""
    if not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect("home")
    
    election = get_object_or_404(Election, id=election_id)
    candidates = election.candidates.annotate(vote_count=Count("votes")).order_by("-vote_count")
    total_votes = election.total_votes()

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
    )
    
    elements.append(Paragraph(f"Election Results: {election.title}", title_style))
    elements.append(Spacer(1, 0.3 * inch))
    
    # Election info
    info_text = f"Total Votes: {total_votes} | Date: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))
    
    # Results table
    table_data = [['Rank', 'Candidate', 'Party', 'Votes', 'Percentage']]
    for idx, candidate in enumerate(candidates, 1):
        percentage = candidate.vote_percentage(total_votes)
        table_data.append([
            str(idx),
            candidate.name,
            candidate.party or 'Independent',
            str(candidate.vote_count),
            f"{percentage}%"
        ])
    
    table = Table(table_data, colWidths=[0.8*inch, 2*inch, 1.5*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5AA0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="election_results_{election_id}.pdf"'
    return response


@require_POST
@login_required
def export_results_excel(request, election_id):
    """Export election results as Excel"""
    if not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect("home")
    
    election = get_object_or_404(Election, id=election_id)
    candidates = election.candidates.annotate(vote_count=Count("votes")).order_by("-vote_count")
    total_votes = election.total_votes()

    # Create Excel file
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Results"
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    
    # Header
    headers = ['Rank', 'Candidate', 'Party', 'Votes', 'Percentage']
    header_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for idx, candidate in enumerate(candidates, 1):
        percentage = candidate.vote_percentage(total_votes)
        worksheet.cell(row=idx+1, column=1).value = idx
        worksheet.cell(row=idx+1, column=2).value = candidate.name
        worksheet.cell(row=idx+1, column=3).value = candidate.party or 'Independent'
        worksheet.cell(row=idx+1, column=4).value = candidate.vote_count
        worksheet.cell(row=idx+1, column=5).value = f"{percentage}%"
    
    # Summary
    summary_row = len(candidates) + 3
    worksheet.cell(row=summary_row, column=1).value = "Total Votes:"
    worksheet.cell(row=summary_row, column=2).value = total_votes
    
    # Save and return
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="election_results_{election_id}.xlsx"'
    return response


# ============ Admin Panel Views ============

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Admin dashboard"""
    total_elections = Election.objects.count()
    active_elections = Election.objects.filter(is_active=True).count()
    total_votes = Vote.objects.count()
    total_users = CustomUser.objects.count()
    verified_voters = VoterProfile.objects.filter(verification_status='verified').count()
    
    recent_votes = list(
        Vote.objects.values(
            'user__username',
            'election_id',
            'election__title',
            'candidate__name',
            'voted_at',
        ).order_by('-voted_at')[:10]
    )
    recent_logs = VoteLog.objects.order_by('-timestamp')[:10]
    
    context = {
        'total_elections': total_elections,
        'active_elections': active_elections,
        'total_votes': total_votes,
        'total_users': total_users,
        'verified_voters': verified_voters,
        'recent_votes': recent_votes,
        'recent_logs': recent_logs,
    }
    return render(request, "voting/admin_dashboard.html", context)


@login_required
@user_passes_test(is_admin)
def manage_elections(request):
    """Manage elections"""
    elections = list(
        Election.objects.annotate(vote_count=Count("votes")).values(
            "id", "title", "description", "vote_count"
        ).order_by("-id")
    )
    return render(request, "voting/manage_elections.html", {"elections": elections})


@login_required
@user_passes_test(is_admin)
def create_election(request):
    """Create new election"""
    if request.method == "POST":
        form = ElectionForm(request.POST)
        if form.is_valid():
            election = Election(
                title=form.cleaned_data['title'],
                description=form.cleaned_data['description'],
                start_time=form.cleaned_data['start_time'],
                end_time=form.cleaned_data['end_time'],
                require_voter_verification=form.cleaned_data['require_voter_verification'],
                publish_results=form.cleaned_data['publish_results'],
                created_by=request.user
            )
            election.save()
            
            # Log action
            AuditLog.objects.create(
                admin_user=request.user,
                action_type='create',
                model_name='Election',
                object_id=election.id,
                description=f'Created election: {election.title}'
            )
            
            messages.success(request, "Election created successfully!")
            return redirect("manage_elections")
    else:
        form = ElectionForm()
    
    return render(request, "voting/create_election.html", {"form": form})


@login_required
@user_passes_test(is_admin)
def manage_candidates(request, election_id):
    """Manage candidates for an election"""
    election = Election.objects.filter(id=election_id).values("id", "title").first()
    if not election:
        raise Http404("Election not found")
    candidates = Candidate.objects.filter(election_id=election_id).order_by("position", "name")
    return render(request, "voting/manage_candidates.html", {
        "election": election,
        "candidates": candidates
    })


@login_required
@user_passes_test(is_admin)
def add_candidate(request, election_id):
    """Add candidate to election"""
    election = get_object_or_404(Election, id=election_id)
    
    if request.method == "POST":
        form = CandidateForm(request.POST, request.FILES)
        if form.is_valid():
            candidate = form.save(commit=False)
            candidate.election = election
            candidate.save()
            
            AuditLog.objects.create(
                admin_user=request.user,
                action_type='create',
                model_name='Candidate',
                object_id=candidate.id,
                description=f'Added candidate: {candidate.name} to {election.title}'
            )
            
            messages.success(request, "Candidate added successfully!")
            return redirect("manage_candidates", election_id=election_id)
    else:
        form = CandidateForm()
    
    return render(request, "voting/add_candidate.html", {
        "election": election,
        "form": form
    })


@login_required
@user_passes_test(is_admin)
def verify_voters(request):
    """Verify voter profiles"""
    pending_verifications = VoterProfile.objects.filter(verification_status='pending')
    
    if request.method == "POST":
        profile_id = request.POST.get('profile_id')
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        profile = get_object_or_404(VoterProfile, id=profile_id)
        
        if action == 'approve':
            profile.verification_status = 'verified'
            profile.verified_by = request.user
            profile.verified_at = timezone.now()
            profile.verification_notes = notes
            profile.save()
            
            # Send notification
            NotificationLog.objects.create(
                recipient=profile.user,
                notification_type='email',
                subject='Voter Profile Verified',
                message='Your voter profile has been verified successfully.',
                status='sent'
            )
            
            messages.success(request, f"Voter {profile.user.username} verified!")
        
        elif action == 'reject':
            profile.verification_status = 'rejected'
            profile.verified_by = request.user
            profile.verified_at = timezone.now()
            profile.verification_notes = notes
            profile.save()
            
            messages.warning(request, f"Voter {profile.user.username} rejected!")
    
    context = {
        'pending_verifications': pending_verifications,
    }
    return render(request, "voting/verify_voters.html", context)


@login_required
def view_vote_history(request):
    """View voting history and activity logs"""
    vote_history = VoteLog.objects.filter(user=request.user).order_by('-timestamp')
    
    context = {
        'vote_history': vote_history,
    }
    return render(request, "voting/vote_history.html", context)


@login_required
def settings_view(request):
    """User settings"""
    if request.method == "POST":
        form = LanguagePreferenceForm(request.POST)
        if form.is_valid():
            request.user.language = form.cleaned_data['language']
            request.user.save()
            messages.success(request, "Language preference updated!")
            return redirect("settings")
    else:
        form = LanguagePreferenceForm(initial={'language': request.user.language})
    
    context = {
        'form': form,
        'two_fa_enabled': request.user.two_fa_enabled,
        'email_verified': request.user.email_verified,
    }
    return render(request, "voting/settings.html", context)


# ============ API Helper Views ============

@login_required
def get_election_stats_json(request, election_id):
    """Get election statistics as JSON"""
    election = get_object_or_404(Election, id=election_id)
    
    if not election.publish_results and not request.user.is_staff:
        return JsonResponse({'error': 'Results not published'}, status=403)
    
    stats = get_election_stats(election)
    
    data = {
        'total_votes': stats['total_votes'],
        'total_voters': stats['total_voters'],
        'candidates': [
            {
                'id': c.id,
                'name': c.name,
                'party': c.party,
                'votes': c.vote_count(),
                'percentage': c.vote_percentage(stats['total_votes']),
            }
            for c in stats['candidates']
        ]
    }
    
    return JsonResponse(data)
